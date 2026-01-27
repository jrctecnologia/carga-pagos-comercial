from datetime import date, datetime
from openpyxl import load_workbook

from config import EXCEL_FILE, SHEET_NAME

def leer_filas_desde_excel():
    """Lee todas las facturas del Excel y retorna una lista de tuplas con todos los datos necesarios"""
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    facturas = []
    # Empezamos desde la fila 3 (asumiendo que 1 y 2 son encabezados)
    for fila in range(3, ws.max_row + 1):
        codigo_cliente = ws[f"A{fila}"].value
        serie_factura = ws[f"D{fila}"].value
        folio_factura = ws[f"E{fila}"].value
        moneda_factura = ws[f"G{fila}"].value
        fecha_pago = ws[f"H{fila}"].value
        serie_pago = ws[f"I{fila}"].value
        folio_pago = ws[f"J{fila}"].value
        importe_pago = ws[f"K{fila}"].value
        impuesto1 = ws[f"L{fila}"].value
        porc_impuesto1 = ws[f"M{fila}"].value
        moneda_pago = ws[f"N{fila}"].value
        tipo_cambio = ws[f"O{fila}"].value
        referencia = ws[f"P{fila}"].value
        
        # Solo procesar si hay datos relevantes
        if codigo_cliente is not None or serie_factura is not None or folio_factura is not None:
            facturas.append((fila, codigo_cliente, serie_factura, folio_factura, moneda_factura, fecha_pago, serie_pago, folio_pago, importe_pago, impuesto1, porc_impuesto1, moneda_pago, tipo_cambio, referencia))
    
    wb.close()
    return facturas

def formatea_fecha_pago(fecha_pago):
    if isinstance(fecha_pago, (datetime, date)):
        return fecha_pago.strftime("%Y%m%d")
    if fecha_pago is None:
        return None
    return str(fecha_pago)


def imprime_resultado(
    archivo_excel,
    hoja,
    fila,
    mensaje,
    columna_resultado="Q"
):
    wb = load_workbook(archivo_excel)
    ws = wb[hoja]

    celda = f"{columna_resultado}{fila}"
    ws[celda] = mensaje

    wb.save(archivo_excel)
    wb.close()
